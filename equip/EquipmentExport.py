# Created by moeheart at 03/29/2021
# 配装导出库，用于与计算器对接，实现一站式操作。

from openpyxl import load_workbook
from equip.EquipmentType import *

def getPlug(id):
    '''
    根据五行石id获取五行石等级.
    params:
    - id: 五行石id
    returns:
    - 五行石等级
    '''
    plugDict = {"": 0, "24423": 1, "24424": 2, "24425": 3, "24426": 4, "24427": 5, "24428": 6, "24429": 7, "24430": 8,
                "24442": 1, "24443": 2, "24444": 3, "24445": 4, "24446": 5, "24447": 6, "24448": 7, "24449": 8}
    if id not in plugDict:
        return 0
    else:
        return plugDict[id]

class HuajianExportEquipment():
    '''
    花间计算器的导出类。
    将json形式的数据导出到计算器excel表中。
    现在的版本会影响excel的完整性，等待以后修复。
    '''
    
    equip6 = """24981	聚散囊·4000品（破防破招）
25014	云愁囊·4000品（破防无双）
25047	帘灭囊·4000品（会心破招）
25098	天颜囊·4340品（会心无双）
25104	残灯囊·4250品（破招无双）
25137	消魂囊·4250品（破防破招）
25170	闲花囊·4250品（加速无双）
25203	静定囊·4250品（破防无双）
25236	歇星·4250品（会心无双）
25242	箕峰囊·4250品（会心破招）
25329	长离箭囊·4620品（加速破招）
25335	百味囊·4620品（会心无双）
25368	飞白囊·4620品（破防破招）
25374	虚冲囊·4620品（会心破招）
25435	建炎箭囊·5070品（加速破招）
25441	流年囊·5070品（会心无双）
25474	玉带囊·5070品（破防破招）
25507	涯风囊·5070品（会心破招）
25594	无皇囊·4000品精简（破防）
25595	无皇囊·4000品精简（会心会效）
25596	无皇囊·4000品精简（会心会效破招）
25602	无皇囊·4160品精简（会心）
25603	无皇囊·4160品精简（破防破招）
25604	无皇囊·4160品精简（会心破防破招）
25610	无皇囊·4320品精简（无双）
25611	无皇囊·4320品精简（会心破招）
25612	无皇囊·4320品精简（会心会效破防）
25618	无皇囊·4480品精简（破防）
25619	无皇囊·4480品精简（会心会效）
25620	无皇囊·4480品精简（会心会效破招）
25626	无皇囊·4640品精简（会心）
25627	无皇囊·4640品精简（破防破招）
25628	无皇囊·4640品精简（会心破防破招）
25634	无皇囊·4800品精简（无双）
25635	无皇囊·4800品精简（会心破招）
25636	无皇囊·4800品精简（会心会效破防）
25642	无皇囊·4960品精简（破防）
25643	无皇囊·4960品精简（会心会效）
25644	无皇囊·4960品精简（会心会效破招）
25948	涸鱼得水·幽静囊·4550品（会心无双）
26935	来年春野·岁 ·4900品（破防无双）
24988	灯又烬·4000品（破防无双）
25021	芳丛·4000品（加速破招）
25054	分流·4000品（会心破招）
25111	月在庭花·4250品（破防无双）
25144	松江·4250品（加速破招）
25177	观玉颜·4250品（会心破招）
25210	寒光·4250品（破防破招）
25249	图箕笔·4250品（会心无双）
25276	寻踪觅宝·阆苑·4250品（会心破招）
25303	寻踪觅宝·早晖·4550品（会心破招）
25342	阔别·4620品（会心无双）
25381	红叶笔·4620品（破防破招）
25448	开落荣枯·5070品（会心无双）
25481	清浦·5070品（破防无双）
25514	客庭寒·5070品（破防破招）
25569	涵流·5140品（会心破招）
25652	闲心·4250品橙武（加速会心破防）
25955	涸鱼得水·白兔·4550品（破防无双）
26680	雪凤冰王笛·5070品（破防破招）
26837	翡翠玲珑筷·火候·4500品（破防无双）
26882	哭佛点睛笔·山水·4850品（破防破招）"""

    equip7 = """50094	聚散帽·4000品（会心破招）
50130	云愁帽·4000品（加速无双）
50166	帘灭冠·4000品（破防破招）
50244	天营帽·4060品（破防无双）
50268	天营帽·4340品（破防无双）
50298	残灯冠·4250品（会心无双）
50334	消魂冠·4250品（加速破招）
50370	闲花冠·4250品（会心破招）
50406	静定冠·4250品（破防无双）
50436	箕峰帽·4250品（破防破招）
50472	知微通玄·幽梦帽·4250品（破防破招）
50502	寻踪觅宝·修阔帽·4250品（会心无双）
50593	星演·池南冠·4550品套装（破防破招）
50616	星演·鸣珂冠·4550品（破防破招）
50668	知微通玄·幽梦帽·4550品（破防破招）
50698	寻踪觅宝·永清帽·4550品（会心无双）
50734	长离冠·4620品（会心无双）
50770	百味帽·4620品（加速破招）
50806	虚冲冠·4620品（会心破招）
50927	惊尘·云涯冠·5000品套装（破防破招）
50950	惊尘·岳列冠·5000品（破防破招）
51002	建炎冠·5070品（会心无双）
51038	流年帽·5070品（加速破招）
51098	涯风冠·5070品（会心破招）
51164	归央冠·5140品（会心破招）
51188	客行江湖·梦桑冠·4620品（会心破招）
51206	客行江湖·雷音冠·5070品（会心破招）
51211	无皇头饰·4000品精简（无双）
51212	无皇头饰·4000品精简（会心无双）
51213	无皇头饰·4000品精简（会心破防破招）
51243	无皇头饰·4160品精简（会心）
51244	无皇头饰·4160品精简（会心破招）
51245	无皇头饰·4160品精简（会心破防无双）
51275	无皇头饰·4320品精简（破防）
51276	无皇头饰·4320品精简（会心会效）
51277	无皇头饰·4320品精简（会心破招无双）
51307	无皇头饰·4480品精简（无双）
51308	无皇头饰·4480品精简（会心无双）
51309	无皇头饰·4480品精简（会心破防破招）
51339	无皇头饰·4640品精简（会心）
51340	无皇头饰·4640品精简（会心破招）
51341	无皇头饰·4640品精简（会心破防无双）
51371	无皇头饰·4800品精简（破防）
51372	无皇头饰·4800品精简（会心会效）
51373	无皇头饰·4800品精简（会心破招无双）
51403	无皇头饰·4960品精简（无双）
51404	无皇头饰·4960品精简（会心无双）
51405	无皇头饰·4960品精简（会心破防破招）
52252	涸鱼得水·幽静冠·4550品（会心无双）
50100	聚散衫·4000品（会心破招）
50136	云愁衫·4000品（加速无双）
50172	帘灭衫·4000品（破防无双）
50304	残灯衣·4250品（破防无双）
50340	消魂衣·4250品（会心破招）
50376	闲花衣·4250品（破防破招）
50412	静定衣·4250品（会心无双）
50442	箕峰衫·4250品（破招无双）
50478	知微通玄·幽梦衣·4250品（加速无双）
50508	寻踪觅宝·修阔衫·4250品（加速破招）
50619	星演·池南衣·4550品套装（加速无双）
50642	星演·鸣珂衫·4550品（加速无双）
50674	知微通玄·幽梦衣·4550品（加速无双）
50704	寻踪觅宝·永清衫·4550品（加速破招）
50740	长离衫·4620品（破防破招）
50776	百味衣·4620品（会心破招）
50812	虚冲衣·4620品（会心无双）
50953	惊尘·云涯衣·5000品套装（加速无双）
50976	惊尘·岳列衣·5000品（加速无双）
51008	建炎衫·5070品（破防破招）
51044	流年衣·5070品（会心破招）
51104	涯风衫·5070品（会心无双）
51170	归央衣·5140品（破防无双）
52258	涸鱼得水·幽静衣·4550品（破防破招）
53689	翡翠御厨上衣·火候·4500品（破防无双）
53712	寻龙御厨上衣·火候·4500品（破防无双）
50076	聚散带·4000品（破招无双）
50112	云愁带·4000品（破防无双）
50148	帘灭带·4000品（加速无双）
50232	天逊腰带·4060品（破防无双）
50256	天逊腰带·4340品（破防无双）
50280	残灯腰带·4250品（会心无双）
50316	消魂腰带·4250品（破防破招）
50352	闲花腰带·4250品（破防无双）
50388	静定腰带·4250品（加速破招）
50454	知微通玄·幽梦带·4250品（会心无双）
50490	寻踪觅宝·修阔带·4250品（加速破招）
50541	星演·池南腰带·4550品套装（破防无双）
50564	星演·鸣珂护腰·4550品（破防无双）
50650	知微通玄·幽梦带·4550品（会心无双）
50686	寻踪觅宝·永清带·4550品（加速破招）
50716	长离带·4620品（破防无双）
50752	百味腰带·4620品（加速破招）
50788	虚冲护腰·4620品（会心破招）
50875	惊尘·云涯腰带·5000品套装（破防无双）
50898	惊尘·岳列腰带·5000品（破防无双）
50984	建炎带·5070品（破防无双）
51020	流年腰带·5070品（加速破招）
51049	倾怀腰带·5070品精简（无双）
51050	重游腰带·5070品精简（破防破招）
51051	烟桥腰带·5070品精简（会心无双）
51055	露华腰带·5070品精简（破防无双）
51080	涯风缠腰·5070品（会心破招）
51146	归央腰带·5140品（会心破招）
52234	涸鱼得水·幽静腰带·4550品（破防破招）
53663	翡翠御厨腰带·火候·4500品（会心破招）
53686	寻龙御厨腰带·火候·4500品（会心破招）
50070	聚散袖·4000品（加速破招）
50106	云愁袖·4000品（破招无双）
50142	帘灭护手·4000品（破防无双）
50226	天源护手·4060品（会心破招）
50250	天源护手·4340品（会心破招）
50274	残灯护手·4250品（加速无双）
50310	消魂护手·4250品（破防破招）
50346	闲花袖·4250品（会心无双）
50382	静定袖·4250品（破防无双）
50418	箕峰袖·4250品（会心无双）
50448	知微通玄·幽梦袖·4250品（破防无双）
50484	寻踪觅宝·修阔袖·4250品（破防破招）
50515	星演·池南护手·4550品套装（会心破招）
50538	星演·鸣珂护手·4550品（会心破招）
50644	知微通玄·幽梦袖·4550品（破防无双）
50680	寻踪觅宝·永清袖·4550品（破防破招）
50710	长离护手·4620品（会心无双）
50746	百味护手·4620品（破防无双）
50782	虚冲护手·4620品（会心破招）
50849	惊尘·云涯护手·5000品套装（会心破招）
50872	惊尘·岳列护手·5000品（会心破招）
50978	建炎护手·5070品（会心无双）
51014	流年护手·5070品（破防无双）
51057	泓碧护腕·5070品精简（破防破招）
51058	玉淙护腕·5070品精简（破防无双）
51059	堂辉护腕·5070品精简（会心无双）
51063	广楼护腕·5070品精简（破招无双）
51074	涯风护手·5070品（会心破招）
51140	归央袖·5140品（破防破招）
51176	客行江湖·梦桑护手·4620品（破防破招）
51194	客行江湖·雷音护手·5070品（破防破招）
51219	无皇护臂·4000品精简（会心）
51220	无皇护臂·4000品精简（会心破防）
51221	无皇护臂·4000品精简（会心会效破防）
51251	无皇护臂·4160品精简（破防）
51252	无皇护臂·4160品精简（会心会效）
51253	无皇护臂·4160品精简（会心会效破招）
51283	无皇护臂·4320品精简（无双）
51284	无皇护臂·4320品精简（破防破招）
51285	无皇护臂·4320品精简（会心会效无双）
51315	无皇护臂·4480品精简（会心）
51316	无皇护臂·4480品精简（会心破防）
51317	无皇护臂·4480品精简（会心会效破防）
51347	无皇护臂·4640品精简（破防）
51348	无皇护臂·4640品精简（会心会效）
51349	无皇护臂·4640品精简（会心会效破招）
51379	无皇护臂·4800品精简（无双）
51380	无皇护臂·4800品精简（破防破招）
51381	无皇护臂·4800品精简（会心会效无双）
51411	无皇护臂·4960品精简（会心）
51412	无皇护臂·4960品精简（会心破防）
51413	无皇护臂·4960品精简（会心会效破防）
52228	涸鱼得水·幽静护手·4550品（会心破招）
50088	聚散裤·4000品（加速破招）
50124	云愁裤·4000品（破防破招）
50160	帘灭裳·4000品（会心无双）
50292	残灯裤·4250品（破防无双）
50328	消魂裤·4250品（加速破招）
50364	闲花裤·4250品（会心破招）
50400	静定裤·4250品（破防破招）
50430	箕峰裤·4250品（会心无双）
50466	知微通玄·幽梦裤·4250品（会心破招）
50662	知微通玄·幽梦裤·4550品（会心破招）
50728	长离下裳·4620品（加速破招）
50764	百味裤·4620品（破招无双）
50800	虚冲裳·4620品（会心破招）
50996	建炎下裳·5070品（加速破招）
51032	流年裤·5070品（破招无双）
51065	吞山裤·5070品精简（会心无双）
51066	披霜裤·5070品精简（破防破招）
51067	荒涧裤·5070品精简（破招无双）
51071	历岁裤·5070品精简（破防无双）
51092	涯风裳·5070品（会心破招）
51158	归央裤·5140品（会心无双）
51182	客行江湖·梦桑下裳·4620品（破防无双）
51200	客行江湖·雷音下裳·5070品（破防无双）
51227	无皇裤·4000品精简（破防）
51228	无皇裤·4000品精简（破招无双）
51229	无皇裤·4000品精简（会心破防无双）
51259	无皇裤·4160品精简（无双）
51260	无皇裤·4160品精简（会心无双）
51261	无皇裤·4160品精简（破防破招无双）
51291	无皇裤·4320品精简（会心）
51292	无皇裤·4320品精简（破防无双）
51293	无皇裤·4320品精简（会心破防破招）
51323	无皇裤·4480品精简（破防）
51324	无皇裤·4480品精简（破招无双）
51325	无皇裤·4480品精简（会心破防无双）
51355	无皇裤·4640品精简（无双）
51356	无皇裤·4640品精简（会心无双）
51357	无皇裤·4640品精简（破防破招无双）
51387	无皇裤·4800品精简（会心）
51388	无皇裤·4800品精简（破防无双）
51389	无皇裤·4800品精简（会心破防破招）
51419	无皇裤·4960品精简（破防）
51420	无皇裤·4960品精简（破招无双）
51421	无皇裤·4960品精简（会心破防无双）
52246	涸鱼得水·幽静裤·4550品（破防破招）
50082	聚散履·4000品（加速无双）
50118	云愁靴·4000品（破防破招）
50154	帘灭履·4000品（会心破招）
50238	天鸿鞋·4060品（会心破招）
50262	天鸿鞋·4340品（会心破招）
50286	残灯靴·4250品（会心无双）
50322	消魂靴·4250品（加速破招）
50358	闲花履·4250品（破防破招）
50394	静定靴·4250品（加速无双）
50424	箕峰靴·4250品（破防无双）
50460	知微通玄·幽梦靴·4250品（加速破招）
50496	寻踪觅宝·修阔靴·4250品（会心无双）
50567	星演·池南履·4550品套装（破防无双）
50590	星演·鸣珂靴·4550品（破防无双）
50656	知微通玄·幽梦靴·4550品（加速破招）
50692	寻踪觅宝·永清靴·4550品（会心无双）
50722	长离靴·4620品（加速破招）
50758	百味靴·4620品（会心无双）
50794	虚冲靴·4620品（会心破招）
50901	惊尘·云涯靴·5000品套装（破防无双）
50924	惊尘·岳列靴·5000品（破防无双）
50990	建炎靴·5070品（加速破招）
51026	流年靴·5070品（会心无双）
51086	涯风靴·5070品（会心破招）
51152	归央靴·5140品（破防破招）
51235	无皇鞋·4000品精简（无双）
51236	无皇鞋·4000品精简（破防无双）
51237	无皇鞋·4000品精简（会心破招无双）
51267	无皇鞋·4160品精简（破防）
51268	无皇鞋·4160品精简（破招无双）
51269	无皇鞋·4160品精简（会心会效无双）
51299	无皇鞋·4320品精简（会心）
51300	无皇鞋·4320品精简（会心破防）
51301	无皇鞋·4320品精简（破防破招无双）
51331	无皇鞋·4480品精简（无双）
51332	无皇鞋·4480品精简（破防无双）
51333	无皇鞋·4480品精简（会心破招无双）
51363	无皇鞋·4640品精简（破防）
51364	无皇鞋·4640品精简（破招无双）
51365	无皇鞋·4640品精简（会心会效无双）
51395	无皇鞋·4800品精简（会心）
51396	无皇鞋·4800品精简（会心破防）
51397	无皇鞋·4800品精简（破防破招无双）
51427	无皇鞋·4960品精简（无双）
51428	无皇鞋·4960品精简（破防无双）
51429	无皇鞋·4960品精简（会心破招无双）
52240	涸鱼得水·幽静靴·4550品（会心无双）"""


    equip8 = """29123	聚散链·4000品（破招无双）
29141	云愁链·4000品（会心破招）
29159	帘灭链·4000品（破防无双）
29213	天首链·4060品（破防破招）
29231	天首链·4340品（破防破招）
29249	残灯链·4250品（加速无双）
29267	消魂链·4250品（破防破招）
29285	闲花链·4250品（会心无双）
29303	静定链·4250品（会心破招）
29321	箕峰链·4250品（破防无双）
29333	长离链·4620品（破防无双）
29351	百味项饰·4620品（会心无双）
29369	虚冲链·4620品（会心破招）
29387	建炎链·5070品（破防无双）
29405	流年项饰·5070品（会心无双）
29437	涯风项饰·5070品（会心破招）
29455	燃香链·4360品（破防破招）
29467	燃香链·4620品（破防破招）
29479	燃香链·4880品（破防破招）
29491	燃香链·5140品（破防破招）
29514	无皇项链·4000品精简（破防）
29515	无皇项链·4000品精简（会心破招）
29516	无皇项链·4000品精简（破防破招无双）
29530	无皇项链·4160品精简（无双）
29531	无皇项链·4160品精简（破防无双）
29532	无皇项链·4160品精简（会心破招无双）
29546	无皇项链·4320品精简（会心）
29547	无皇项链·4320品精简（破招无双）
29548	无皇项链·4320品精简（会心会效破招）
29562	无皇项链·4480品精简（破防）
29563	无皇项链·4480品精简（会心破招）
29564	无皇项链·4480品精简（破防破招无双）
29578	无皇项链·4640品精简（无双）
29579	无皇项链·4640品精简（破防无双）
29580	无皇项链·4640品精简（会心破招无双）
29594	无皇项链·4800品精简（会心）
29595	无皇项链·4800品精简（破招无双）
29596	无皇项链·4800品精简（会心会效破招）
29610	无皇项链·4960品精简（破防）
29611	无皇项链·4960品精简（会心破招）
29612	无皇项链·4960品精简（破防破招无双）
29129	聚散佩·4000品（破招无双）
29147	云愁佩·4000品（加速破招）
29165	帘灭坠·4000品（会心无双）
29219	天许坠·4060品（会心破招）
29237	天许坠·4340品（会心破招）
29255	残灯坠·4250品（加速无双）
29273	消魂坠·4250品（会心破招）
29291	闲花坠·4250品（破防无双）
29309	静定坠·4250品（会心无双）
29339	长离坠·4620品（破防破招）
29357	百味坠·4620品（会心破招）
29375	虚冲佩·4620品（会心破招）
29393	建炎坠·5070品（破防破招）
29411	流年坠·5070品（会心破招）
29423	游仙·5070品特效（破防无双）
29443	涯风佩·5070品（会心破招）
29461	燃香坠·4360品（会心无双）
29473	燃香坠·4620品（会心无双）
29485	燃香坠·4880品（会心无双）
29497	燃香坠·5140品（会心无双）
29522	无皇腰坠·4000品精简（无双）
29523	无皇腰坠·4000品精简（破防破招）
29524	无皇腰坠·4000品精简（会心会效无双）
29538	无皇腰坠·4160品精简（会心）
29539	无皇腰坠·4160品精简（会心破防）
29540	无皇腰坠·4160品精简（会心会效破防）
29554	无皇腰坠·4320品精简（破防）
29555	无皇腰坠·4320品精简（会心无双）
29556	无皇腰坠·4320品精简（会心破防无双）
29570	无皇腰坠·4480品精简（无双）
29571	无皇腰坠·4480品精简（破防破招）
29572	无皇腰坠·4480品精简（会心会效无双）
29586	无皇腰坠·4640品精简（会心）
29587	无皇腰坠·4640品精简（会心破防）
29588	无皇腰坠·4640品精简（会心会效破防）
29602	无皇腰坠·4800品精简（破防）
29603	无皇腰坠·4800品精简（会心无双）
29604	无皇腰坠·4800品精简（会心破防无双）
29618	无皇腰坠·4960品精简（无双）
29619	无皇腰坠·4960品精简（破防破招）
29620	无皇腰坠·4960品精简（会心会效无双）
29135	聚散戒·4000品（破防破招）
29153	云愁戒·4000品（加速无双）
29171	帘灭戒·4000品（会心破招）
29225	天晚戒·4060品（会心无双）
29243	天晚戒·4340品（会心无双）
29261	残灯戒·4250品（破招无双）
29279	消魂戒·4250品（加速破招）
29297	闲花戒·4250品（会心破招）
29315	静定戒·4250品（破防破招）
29327	箕峰戒·4250品（破防无双）
29345	长离指环·4620品（破防破招）
29363	百味戒·4620品（会心无双）
29381	虚冲戒·4620品（会心破招）
29399	建炎指环·5070品（破防破招）
29417	流年戒·5070品（会心无双）
29428	腾霄戒·5070品精简（无双）
29429	无扉戒·5070品精简（破防破招）
29430	紫金戒·5070品精简（会心无双）
29434	云飞戒·5070品精简（破防无双）
29449	涯风指环·5070品（会心破招）
29503	客行江湖·梦桑戒·4620品（破防无双）
29509	客行江湖·雷音戒·5070品（破防无双）
30722	穷处所·4750品（破防破招）
30729	翡翠御厨戒指·火候·4500品（会心无双）
30752	寻龙御厨戒指·火候·4500品（会心无双）"""

    magic1 = """11025	内功攻击提升175点
11027	破招提升325点
11028	加速提升325点
11048	内功攻击提升155点
11050	破招提升289点
11051	加速提升289点
11167	帽子加速等级提高231
11172	帽子内功攻击提高124
11084	体质提升120点
11085	最大气血值提升1625点
11107	体质提升107点
11108	最大气血值提升1444点
11189	上衣无双等级提高174
11029	体质提升120点
11030	最大气血值提升1625点
11052	体质提升107点
11053	最大气血值提升1444点
11191	腰带无双等级提高174
11071	元气提升73点
11076	内功破防等级提升325点
11077	全会心提升325点
11078	无双等级提升325点
11094	元气提升65点
11099	内功破防等级提升289点
11100	全会心提升289点
11101	无双等级提升289点
11166	护腕全会心等级提高231
11175	护腕内功破防等级提高231
11176	护腕无双等级提高231
11016	元气提升73点
11021	内功破防等级提升325点
11022	全会心提升325点
11023	无双等级提升325点
11039	元气提升65点
11044	内功破防等级提升289点
11045	全会心提升289点
11046	无双等级提升289点
11165	下装全会心等级提高231
11169	下装内功破防等级提高231
11170	下装无双等级提高231
11080	内功攻击提升175点
11082	破招提升325点
11083	加速提升325点
11103	内功攻击提升155点
11105	破招提升289点
11106	加速提升289点
11185	鞋子加速等级提高231
11186	鞋子破招等级提高231
11188	鞋子内功攻击提高124
11160	元气提升73点
11163	内功破防等级提升325点
11164	加速提升325点
11011	内功攻击提升175点
11062	内功攻击提升155点
11153	元气提升73点
11156	内功攻击提升175点
11157	破招等级提升325点
11147	体质提升120点"""

    def initDict(self):
        self.equipf = {}
        self.magicf = {}

        e6 = self.equip6.split('\n')
        for p in e6:
            q = p.split('\t')
            self.equipf["6,%s"%q[0]] = q[1]
            
        e7 = self.equip7.split('\n')
        for p in e7:
            q = p.split('\t')
            self.equipf["7,%s"%q[0]] = q[1]
            
        e8 = self.equip8.split('\n')
        for p in e8:
            q = p.split('\t')
            self.equipf["8,%s"%q[0]] = q[1]
            
        m1 = self.magic1.split('\n')
        for p in m1:
            q = p.split('\t')
            self.magicf[q[0]] = q[1]
        
    def getEquipName(self, id_cat, id):
        s = "%s,%s"%(id_cat, id)
        if s in self.equipf:
            return self.equipf[s]
        else:
            return "未知"
        
    def getMagicName(self, id):
        s = str(id)
        if s in self.magicf:
            return self.magicf[s]
        else:
            return "无"
    
    def export(self, equips):
        exPos = {"2": 13,
                 "3": 4, 
                 "4": 3,
                 "5": 9,
                 "6": 11, 
                 "7": 12, 
                 "8": 5,
                 "9": 10,
                 "10": 7,
                 "11": 8,
                 "12": 6,
                 "0": 14}
                 
        wb = load_workbook('花间DPS配装计算器—奉天证道beta1.01.xlsx')
        ws = wb["花间配装计算器"]

        for pos in exPos.keys():
            row = exPos[pos]
            equip = equips[pos]
            
            name = self.getEquipName(equip["id_cat"], equip["id"])
            ws["F%d"%row] = name
            
            star = equip["star"]
            ws["I%d"%row] = star
            
            magicName = self.getMagicName(equip["magic1"])
            ws["J%d"%row] = magicName
            
            if "plug1" in equip:
                plug1 = equip["plug1"]
                ws["M%d"%row] = plug1
            
            if "plug2" in equip:
                plug2 = equip["plug2"]
                ws["O%d"%row] = plug2

            if "plug3" in equip:
                plug3 = equip["plug3"]
                ws["Q%d"%row] = plug3
                
        ws["C8"] = "记得"
        ws["C9"] = "把" 
        ws["C10"] = "五彩石"
        ws["C11"] = "改了"

        wb.save('计算器手动缝合版.xlsx')

    def __init__(self):
        self.initDict()
        
class ExcelExportEquipment():
    '''
    装备导出类，将json格式转换为https://www.jx3box.com/bbs/22011定义的格式.
    '''
    def export(self, equips):
        '''
        转换.
        params:
        - equips: json形式的装备统计
        returns:
        - result: str形式的装备统计
        '''
        result = ""
        for id in ["2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12", "0", "1"]:
            if id in equips:
                line = equips[id]
                result += line["id"]
                result += "\t"
                if "star" in line:
                    result += str(line["star"])
                else:
                    result += "0"
                result += "\t"
                if "magic1" in line:
                    result += line["magic1"]
                else:
                    result += ""
                result += "\t"
                if "magic2" in line:
                    result += line["magic2"]
                else:
                    result += ""
                result += "\t"
                if "plug1" in line:
                    result += str(line["plug1"])
                else:
                    result += ""
                result += "\t"
                if "plug2" in line:
                    result += str(line["plug2"])
                else:
                    result += ""
                result += "\t"
                if "plug3" in line:
                    result += str(line["plug3"])
                else:
                    result += ""
                result += "\t"
                if "plug0" in line:
                    result += str(line["plug0"])
                else:
                    result += ""
            result += "\n"
            
        return result[:-1].strip('\n')
    
    def __init__(self):
        equips = {}


class ImportExcelEquipment():
    '''
    装备导入类，将由https://www.jx3box.com/bbs/22011定义的格式转换为json格式.
    '''
    equips = {}
    def importData(self, attrStr):
        '''
        转换.
        params:
        - attrStr: str形式的装备统计
        returns:
        - equips: json形式的装备统计
        '''
        res = attrStr.split('\n')
        equips = {}
        for i in range(len(res)):
            line = res[i]
            single = line.split('\t')
            equip = {}
            for j in range(len(self.attribOrder)):
                equip[self.attribOrder[j]] = single[j]
            equip["pos"] = self.orderTable[i]
            equip["id_cat"] = self.scemeTable[i]
            equip["id_full"] = "%s,%s"%(equip["id_cat"], equip["id"])
            equips[equip["pos"]] = equip
        return equips

    def __init__(self):
        self.orderTable = ["2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12", "0", "1"]
        self.attribOrder = ["id", "star", "magic1", "magic2", "plug1", "plug2", "plug3", "plug0"]
        self.scemeTable = ["8", "7", "7", "6", "6", "6", "7", "6", "7", "7", "7", "8", "8"]

class EquipmentAnalyser():
    '''
    装备分析类，将茗伊复盘中的数据转为json形式。
    '''

    def getForge(self, equips):
        '''
        根据全身装备，得到装备的打造信息（精炼、镶嵌、五彩石、小附魔、大附魔）
        '''
        refine = 0
        plug = [0] * 9
        magic1 = [0] * 13
        magic2 = [0] * 13
        plugColor = 0

        for key in equips:
            if key not in ["score", "description", "sketch"]:
                equip = equips[key]
                id_full = "%s,%s"%(equip["id_cat"], equip["id"])

                # 获取精炼部分
                if id_full in EQUIPMENT_TYPE:
                    t = EQUIPMENT_TYPE[id_full]
                else:
                    t = ""
                refine_max = 6  # 暂时使用简易的装备类别来统计，后续可以更换记录方式。
                if t in ["精简", "特效腰坠", "特效武器"]:
                    refine_max = 3
                if t in ["大橙武"]:
                    refine_max = 8
                if equip["star"] >= refine_max:
                    refine += 1

                # 获取镶嵌部分
                for i in [1,2,3]:
                    p = "plug%d"%i
                    if p in equip:
                        plug[equip[p]] += 1
                    else:
                        plug[0] += 1

                # 获取五彩石
                if "plug0" in equip:
                    colorID = equip["plug0"]
                    if colorID in COLOR_TYPE:
                        plugColor = COLOR_TYPE[colorID]
                    else:
                        plugColor = 0

                id = int(key)
                # 获取小附魔
                if equip["magic1"] not in [0, "0", ""]:
                    if equip["magic1"] in ENCHANT_TYPE and ENCHANT_TYPE[equip["magic1"]] == 2:
                        magic1[id] = 2
                    else:
                        magic1[id] = 1

                if equip["magic2"] not in [0, "0", ""]:
                    magic2[id] = 1
                    if equip["magic2"] in ENCHANT_TYPE and ENCHANT_TYPE[equip["magic2"]] == 12:
                        magic2[id] = 2

        magic1Sum2 = 0
        magic1Sum1 = 0
        for i in [2,4,6,7,10,11,12,0]:
            if magic1[i] == 2:
                magic1Sum2 += 1
            elif magic1[i] == 1:
                magic1Sum1 += 1

        magic2Str = "%d%d%d%d%d%d"%(magic2[12], magic2[8], magic2[11], magic2[4], magic2[3], magic2[10])

        result = "%d/%d-%d-%d/%d/%d-%d/%s"%(refine, plug[8], plug[7], plug[6], plugColor, magic1Sum2, magic1Sum1, magic2Str)
        return result
    
    def getSketch(self, equips):
        '''
        根据全身装备，得到简要的装备说明（如4惊尘，2精简，大橙武）
        '''
        res = {"星演": 0, "惊尘": 0, "百相": 0, "择芳": 0, "切糕": 0, "精简": 0, "特效腰坠": 0, "特效武器": 0, "门派特效": 0, "大橙武": 0}
        
        for key in equips:
            if key not in ["score", "description", "sketch"]:
                equip = equips[key]
                id_full = "%s,%s"%(equip["id_cat"], equip["id"])
                if id_full in EQUIPMENT_TYPE:
                    t = EQUIPMENT_TYPE[id_full]
                    if t not in res:
                        res[t] = 1
                    else:
                        res[t] += 1
        
        sketch = []
        #计算套装
        if res["星演"] >= 4 or res["惊尘"] >= 4 or res["百相"] >= 4 or res["择芳"] >= 4:
            sketch.append("4件套")
        elif (res["择芳"] >= 2 or res["惊尘"] >= 2) and (res["星演"] >= 2 or res["百相"] >= 2):
            sketch.append("4件套")
        elif res["惊尘"] >= 2:
            sketch.append("2惊尘")
        elif res["星演"] >= 2:
            sketch.append("2星演")
        elif res["百相"] >= 2:
            sketch.append("2百相")
        elif res["择芳"] >= 2:
            sketch.append("2择芳")
            
        if res["切糕"] >= 1:
            sketch.append("%d切糕"%res["切糕"])
        if res["精简"] >= 1:
            sketch.append("%d精简"%res["精简"])
        if res["特效腰坠"] >= 1:
            sketch.append("腰坠")
        if res["特效武器"] >= 1:
            sketch.append("特效武器")
        if res["门派特效"] >= 1:
            sketch.append("门派特效")
        if res["大橙武"] >= 1:
            sketch.append("大橙武")
            
        sketchStr = ','.join(sketch)
        return sketchStr
        
    def getPlug(self, id):
        plugDict = {"": 0, "24423": 1, "24424": 2, "24425": 3, "24426": 4, "24427": 5, "24428": 6, "24429": 7, "24430": 8,
                    "24442": 1, "24443": 2, "24444": 3, "24445": 4, "24446": 5, "24447": 6, "24448": 7, "24449": 8}
        if id not in plugDict:
            return 0
        else:
            return plugDict[id]

    def convert2(self, s, score=0):
        '''
        使用新版本的数据（json形式的luatable）进行转换.
        params:
        - s 茗伊复盘数据, jcl或jx3dat均可
        - score 装分
        '''
        equips = {}

        equips["score"] = score
        equips["description"] = ""

        for c in s:
            d = s[c]
            equip = {}
            equip["pos"] = d["1"]
            equip["id_cat"] = d["2"]
            equip["id"] = d["3"]
            equip["star"] = int(d["4"])
            for i in range(1, 4):
                plugPos = str(i)
                if plugPos in d["5"]:
                    f = d["5"][plugPos]
                    plugID = f["2"]
                    equip["plug%d" % i] = self.getPlug(plugID)
            if "0" in d["5"]:
                equip["plug0"] = d["5"]["0"]["2"]
            equip["magic1"] = d["6"]
            equip["magic2"] = d["7"]
            equips[equip["pos"]] = equip
        equips["sketch"] = self.getSketch(equips)
        equips["forge"] = self.getForge(equips)
        return equips
    
    # def convert(self, s):
    #     '''
    #     进行转换。使用老格式，会在未来移除。
    #     params
    #     - s 茗伊复盘数据（处在[18]）
    #     '''
    #     a = s[0]['']
    #     b = a[2]['']
    #
    #     equips = {}
    #
    #     equips["score"] = int(a[1])
    #     equips["description"] = ""
    #
    #     for c in b:
    #         d = c['']
    #         equip = {}
    #         equip["pos"] = d[0]
    #         equip["id_cat"] = d[1]
    #         equip["id"] = d[2]
    #         equip["star"] = int(d[3])
    #         if "" in d[4]:
    #             f = d[4][""]
    #             i = 0
    #             for g in f:
    #                 i += 1
    #                 plugID = g[""][1]
    #                 equip["plug%d"%i] = self.getPlug(plugID)
    #         elif "2" in d[4]:
    #             f = d[4]["2"]
    #             i = 1
    #             for g in f:
    #                 i += 1
    #                 plugID = g[""][1]
    #                 equip["plug%d"%i] = self.getPlug(plugID)
    #         elif "3" in d[4]:
    #             f = d[4]["3"]
    #             i = 2
    #             for g in f:
    #                 i += 1
    #                 plugID = g[""][1]
    #                 equip["plug%d"%i] = self.getPlug(plugID)
    #         if "0" in d[4] or "1" in d[4]:
    #             equip["plug0"] = d[4]["0"][0][""][1]
    #         equip["magic1"] = d[5]
    #         equip["magic2"] = d[6]
    #         equips[equip["pos"]] = equip
    #
    #     equips["sketch"] = self.getSketch(equips)
    #
    #     return equips
    
    def __init__(self):
        pass